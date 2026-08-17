"""
pages/2_5_Staffing_Planner.py - Kitchen staffing planner based on role templates.

This page sits between Shifts and Payroll and provides weekly role-based staffing
suggestions using existing employee availability and fallback rules.
"""

import json
import os
import sys
from io import BytesIO
from datetime import date, datetime, time, timedelta
from typing import Optional

import openpyxl
import pandas as pd
import streamlit as st

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database import get_db, init_db
from models import Employee, Shift
from rules import calc_hours

st.set_page_config(page_title="人力配置建議", page_icon="🧩", layout="wide")

init_db()

st.title("🧩 人力配置建議")
st.caption("依既有後廚排班規律自動產生建議名單，可作為 Shifts 頁面的排班草稿。")
st.markdown("---")

_DAY_KEYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
_DAY_LABELS = ["週一", "週二", "週三", "週四", "週五", "週六", "週日"]

# Role backup rules derived from current operation patterns.
PRECOOK_BACKUPS = {"Ari"}
DISHWASHER_BACKUPS = {"VISMAR MARTINEZ(Kevin)"}

DEFAULT_TEMPLATE = {
    "Mon": {"Cook": 2, "Precook": 2, "Dishwasher": 0},
    "Tue": {"Cook": 2, "Precook": 2, "Dishwasher": 1},
    "Wed": {"Cook": 2, "Precook": 2, "Dishwasher": 1},
    "Thu": {"Cook": 3, "Precook": 2, "Dishwasher": 0},
    "Fri": {"Cook": 3, "Precook": 1, "Dishwasher": 1},
    "Sat": {"Cook": 2, "Precook": 2, "Dishwasher": 1},
    "Sun": {"Cook": 2, "Precook": 2, "Dishwasher": 1},
}


# Week navigation
if "staff_week_offset" not in st.session_state:
    st.session_state.staff_week_offset = 0

nav1, nav2, nav3 = st.columns([1, 2, 1])
with nav1:
    if st.button("← 上週", key="staff_prev"):
        st.session_state.staff_week_offset -= 1
        st.rerun()
with nav3:
    if st.button("下週 →", key="staff_next"):
        st.session_state.staff_week_offset += 1
        st.rerun()
with nav2:
    if st.button("🏠 回本週", key="staff_now"):
        st.session_state.staff_week_offset = 0
        st.rerun()

base_monday = date.today() - timedelta(days=date.today().weekday())
week_start = base_monday + timedelta(weeks=st.session_state.staff_week_offset)
week_days = [week_start + timedelta(days=i) for i in range(7)]

st.subheader(f"規劃週期：{week_days[0].strftime('%Y/%m/%d')}（週一）～ {week_days[-1].strftime('%Y/%m/%d')}（週日）")


@st.cache_data(show_spinner=False)
def parse_availability(raw: str) -> dict:
    if not raw or raw in ("{}", "null"):
        return {}
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return {}


def is_available_for_window(emp: Employee, day_key: str, start_s: str, end_s: str) -> bool:
    av = parse_availability(emp.availability)
    if not av:
        return True
    day_slots = set(av.get(day_key, []))
    if not day_slots:
        return False

    start_t = datetime.strptime(start_s, "%H:%M").time()
    end_t = datetime.strptime(end_s, "%H:%M").time()

    cur = datetime.combine(date.today(), start_t)
    stop = datetime.combine(date.today(), end_t)
    if stop <= cur:
        stop += timedelta(days=1)

    while cur < stop:
        if cur.strftime("%H:%M") not in day_slots:
            return False
        cur += timedelta(minutes=30)
    return True


def weekly_existing_hours(shifts: list[Shift]) -> dict[int, float]:
    h: dict[int, float] = {}
    for s in shifts:
        h[s.employee_id] = h.get(s.employee_id, 0.0) + calc_hours(s.start_time, s.end_time, s.break_minutes)
    return h


def candidate_source(emp: Employee, role: str) -> str:
    if emp.position == role:
        return "primary"
    if role == "Precook" and emp.name in PRECOOK_BACKUPS:
        return "backup"
    if role == "Dishwasher" and emp.name in DISHWASHER_BACKUPS:
        return "backup"
    return "cross"


def role_candidates(employees: list[Employee], role: str, day_key: str, start_s: str, end_s: str) -> list[Employee]:
    result = []
    for emp in employees:
        if not is_available_for_window(emp, day_key, start_s, end_s):
            continue
        if emp.position == role:
            result.append(emp)
            continue
        if role == "Precook" and emp.name in PRECOOK_BACKUPS:
            result.append(emp)
            continue
        if role == "Dishwasher" and emp.name in DISHWASHER_BACKUPS:
            result.append(emp)
            continue
    return result


def pick_best(cands: list[Employee], assigned_today: set[int], hours_map: dict[int, float], target_map: dict[int, float], planned_days: dict[int, int]) -> Optional[Employee]:
    pool = [c for c in cands if c.id not in assigned_today]
    if not pool:
        return None

    def score(emp: Employee):
        target = max(1.0, target_map.get(emp.id, 40.0))
        load_ratio = hours_map.get(emp.id, 0.0) / target
        return (
            round(load_ratio, 4),
            planned_days.get(emp.id, 0),
            hours_map.get(emp.id, 0.0),
            emp.name,
        )

    pool.sort(key=score)
    return pool[0]


def build_boh_workbook_bytes(
    recommendation_rows: list[dict],
    week_dates: list[date],
    employees_in_scope: list[Employee],
) -> bytes:
    """Build a BOH-style workbook and auto-stagger breaks with max 3 concurrent breaks."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOH Schedule"

    day_keys = [d.strftime("%Y-%m-%d") for d in week_dates]
    day_labels = ["MON", "TUES", "WED", "THURS", "FRI", "SAT", "SUN"]

    # Header rows (same layout as source BOH sheet)
    ws.cell(1, 1, f"WEEKLY SCHEDULE - {week_dates[0].strftime('%m/%d/%Y')} to {week_dates[-1].strftime('%m/%d/%Y')}")
    ws.cell(2, 1, "Name")
    for i, d in enumerate(week_dates):
        base_col = 2 + i * 2
        ws.cell(2, base_col, d.strftime("%m/%d"))
        ws.cell(3, base_col, day_labels[i])
        ws.cell(4, base_col, "Clock In")
        ws.cell(4, base_col + 1, "Clock Out")

    def hhmm_to_min(s: str) -> int:
        dt = datetime.strptime(s.strip(), "%H:%M")
        return dt.hour * 60 + dt.minute

    def min_to_time(m: int) -> time:
        m = m % (24 * 60)
        return time(m // 60, m % 60)

    def shift_str_to_window(shift_text: str) -> tuple[int, int]:
        start_s, end_s = shift_text.split("-", 1)
        s = hhmm_to_min(start_s)
        e = hhmm_to_min(end_s)
        if e <= s:
            e += 24 * 60
        return s, e

    break_durations = [10, 60, 30, 10]
    break_labels = ["10 Mins Break", "60 Mins Meal", "30 Mins Meal", "10 Mins Break"]
    # Relative to shift start; matches current BOH style as the baseline.
    break_base_offsets = [150, 210, 390, 540]
    max_concurrent_breaks = 3

    # Build quick lookup and day groups from generated recommendation rows.
    assigned = {(r["人員"], r["日期"]): r for r in recommendation_rows}
    by_day: dict[str, list[dict]] = {}
    for r in recommendation_rows:
        by_day.setdefault(r["日期"], []).append(r)

    # Auto-stagger planner: choose each break start with a cost function and occupancy cap.
    day_break_plan: dict[tuple[str, str], list[int]] = {}
    for day_str, day_rows in by_day.items():
        # Minute-level occupancy on 10-minute grid.
        occupancy: dict[int, int] = {}

        # Stable order keeps output deterministic and close to kitchen priorities.
        role_rank = {"Cook": 1, "Dishwasher": 2, "Precook": 3, "Prep Cook": 3}
        day_rows_sorted = sorted(
            day_rows,
            key=lambda x: (role_rank.get(str(x.get("職務", "")), 9), str(x.get("人員", ""))),
        )

        for rr in day_rows_sorted:
            name = str(rr["人員"])
            shift_s, shift_e = shift_str_to_window(str(rr["班次"]))
            chosen: list[int] = []

            for idx, dur in enumerate(break_durations):
                base = shift_s + break_base_offsets[idx]

                prev_end = chosen[idx - 1] + break_durations[idx - 1] if idx > 0 else shift_s
                earliest = max(prev_end + 30, shift_s + 90 if idx == 0 else prev_end + 30)

                rem_dur = sum(break_durations[idx + 1:])
                rem_work_gap = 30 * (len(break_durations) - idx - 1)
                latest = shift_e - (dur + rem_dur + rem_work_gap + 30)

                candidates = []
                for delta in [0, -10, 10, -20, 20, -30, 30, -40, 40, -50, 50, -60, 60]:
                    cand = base + delta
                    if cand < earliest or cand > latest:
                        continue
                    if cand % 10 != 0:
                        cand = cand - (cand % 10)
                    candidates.append(cand)

                if not candidates:
                    fallback = max(earliest, min(base, latest))
                    fallback = fallback - (fallback % 10)
                    candidates = [fallback]

                def candidate_cost(cand_start: int):
                    peak_after = 0
                    load_sum = 0
                    for m in range(cand_start, cand_start + dur, 10):
                        c = occupancy.get(m, 0)
                        load_sum += c
                        peak_after = max(peak_after, c + 1)
                    over_cap = max(0, peak_after - max_concurrent_breaks)
                    return (over_cap, peak_after, load_sum, abs(cand_start - base))

                best = min(candidates, key=candidate_cost)
                chosen.append(best)
                for m in range(best, best + dur, 10):
                    occupancy[m] = occupancy.get(m, 0) + 1

            day_break_plan[(day_str, name)] = chosen

    # Keep display order similar to kitchen operation: Cook -> Dishwasher -> Precook.
    pos_order = {"Cook": 1, "Dishwasher": 2, "Precook": 3, "Prep Cook": 3}
    employees_sorted = sorted(
        employees_in_scope,
        key=lambda e: (pos_order.get((e.position or "").strip(), 9), e.name),
    )

    row_ptr = 5
    for emp in employees_sorted:
        ws.cell(row_ptr, 1, emp.name)
        ws.cell(row_ptr + 1, 1, (emp.position or "").strip())

        for di, day_str in enumerate(day_keys):
            if (emp.name, day_str) not in assigned:
                continue
            ci_col = 2 + di * 2
            co_col = ci_col + 1

            shift_s, shift_e = shift_str_to_window(str(assigned[(emp.name, day_str)]["班次"]))
            break_starts = day_break_plan.get((day_str, emp.name))
            if not break_starts:
                break_starts = [shift_s + off for off in break_base_offsets]

            b1, b2, b3, b4 = break_starts
            d1, d2, d3, d4 = break_durations

            segments = [
                (min_to_time(shift_s), min_to_time(b1), None),
                (None, None, break_labels[0]),
                (min_to_time(b1 + d1), min_to_time(b2), None),
                (None, None, break_labels[1]),
                (min_to_time(b2 + d2), min_to_time(b3), None),
                (None, None, break_labels[2]),
                (min_to_time(b3 + d3), min_to_time(b4), None),
                (None, None, break_labels[3]),
                (min_to_time(b4 + d4), min_to_time(shift_e), None),
            ]

            for seg_idx, (t_in, t_out, label) in enumerate(segments):
                rr = row_ptr + seg_idx
                if label:
                    ws.cell(rr, ci_col, label)
                else:
                    ws.cell(rr, ci_col, t_in)
                    ws.cell(rr, co_col, t_out)

        row_ptr += len(segments) + 1

    ws.column_dimensions["A"].width = 34
    for col in range(2, 16):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 11

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Load data from DB
db = get_db()
try:
    employees = db.query(Employee).order_by(Employee.name).all()
    shifts = (
        db.query(Shift)
        .filter(Shift.date >= week_days[0], Shift.date <= week_days[-1])
        .all()
    )
finally:
    db.close()

if not employees:
    st.warning("⚠️ 目前沒有員工資料，請先到 Employees 頁面新增員工。")
    st.stop()

# Planner scope
st.markdown("### 1) 規劃範圍")
scope_col1, scope_col2 = st.columns(2)

all_depts = sorted({(e.department or "").strip() for e in employees})
default_dept = "後廚" if "後廚" in all_depts else (all_depts[0] if all_depts else "")
selected_dept = scope_col1.selectbox("部門", all_depts, index=all_depts.index(default_dept) if default_dept in all_depts else 0)
only_active = scope_col2.checkbox("僅顯示本週有可排時段的人員", value=True)

pool = [e for e in employees if (e.department or "").strip() == selected_dept]
if only_active:
    tmp = []
    for e in pool:
        av = parse_availability(e.availability)
        if not av:
            tmp.append(e)
            continue
        if any(av.get(k, []) for k in _DAY_KEYS):
            tmp.append(e)
    pool = tmp

if not pool:
    st.info("目前範圍內沒有可用員工。")
    st.stop()

st.caption(f"規劃人池：{len(pool)} 人")

# Template editor
st.markdown("### 2) 每日職務模板")
default_rows = []
for i, day_key in enumerate(_DAY_KEYS):
    tpl = DEFAULT_TEMPLATE[day_key]
    default_rows.append(
        {
            "星期": _DAY_LABELS[i],
            "日期": week_days[i].strftime("%Y-%m-%d"),
            "Cook": tpl["Cook"],
            "Precook": tpl["Precook"],
            "Dishwasher": tpl["Dishwasher"],
            "班次開始": "11:00",
            "班次結束": "22:00",
            "休息(分)": 30,
        }
    )

plan_df = st.data_editor(
    pd.DataFrame(default_rows),
    use_container_width=True,
    hide_index=True,
    column_config={
        "Cook": st.column_config.NumberColumn(min_value=0, step=1),
        "Precook": st.column_config.NumberColumn(min_value=0, step=1),
        "Dishwasher": st.column_config.NumberColumn(min_value=0, step=1),
        "休息(分)": st.column_config.NumberColumn(min_value=0, max_value=180, step=5),
    },
    key="staff_template_editor",
)

st.caption("備註：Precook 會優先用正職，Ari 可作補位；Dishwasher 會優先用正職，Kevin 可作補位。")

# Generate recommendation
if st.button("🧠 產生本週建議名單", type="primary"):
    existing_hours = weekly_existing_hours(shifts)
    target_hours = {e.id: float(e.target_hours or 40.0) for e in pool}
    planned_days: dict[int, int] = {}

    rows = []
    shortage_rows = []

    for i, day_key in enumerate(_DAY_KEYS):
        day_date = week_days[i]
        day_cfg = plan_df.iloc[i]
        start_s = str(day_cfg["班次開始"])
        end_s = str(day_cfg["班次結束"])
        break_min = int(day_cfg["休息(分)"])

        assigned_today: set[int] = set()
        for role in ["Cook", "Precook", "Dishwasher"]:
            need = int(day_cfg[role])
            cands = role_candidates(pool, role, day_key, start_s, end_s)
            for slot_i in range(need):
                picked = pick_best(cands, assigned_today, existing_hours, target_hours, planned_days)
                if picked is None:
                    shortage_rows.append(
                        {
                            "日期": day_date.strftime("%Y-%m-%d"),
                            "星期": _DAY_LABELS[i],
                            "職務": role,
                            "缺口": 1,
                        }
                    )
                    continue

                assigned_today.add(picked.id)
                planned_days[picked.id] = planned_days.get(picked.id, 0) + 1
                shift_h = calc_hours(
                    datetime.strptime(start_s, "%H:%M").time(),
                    datetime.strptime(end_s, "%H:%M").time(),
                    break_min,
                )
                existing_hours[picked.id] = existing_hours.get(picked.id, 0.0) + shift_h

                rows.append(
                    {
                        "日期": day_date.strftime("%Y-%m-%d"),
                        "星期": _DAY_LABELS[i],
                        "職務": role,
                        "人員": picked.name,
                        "來源": candidate_source(picked, role),
                        "班次": f"{start_s}-{end_s}",
                        "休息(分)": break_min,
                        "本週已排(含建議)": round(existing_hours.get(picked.id, 0.0), 2),
                        "目標工時": round(target_hours.get(picked.id, 40.0), 1),
                    }
                )

    st.markdown("---")
    st.markdown("### 3) 建議排班結果")

    if rows:
        out_df = pd.DataFrame(rows)
        st.dataframe(out_df, use_container_width=True, hide_index=True)

        st.markdown("#### 每日覆蓋摘要")
        daily_summary = (
            out_df.groupby(["日期", "星期", "職務"]).size().reset_index(name="人數")
            .pivot(index=["日期", "星期"], columns="職務", values="人數")
            .fillna(0)
            .reset_index()
        )
        st.dataframe(daily_summary, use_container_width=True, hide_index=True)

        boh_bytes = build_boh_workbook_bytes(rows, week_days, pool)
        boh_name = f"{week_days[0].strftime('%m-%d')}_to_{week_days[-1].strftime('%m-%d')}BOH_v2.xlsx"
        st.download_button(
            label="📥 下載 BOH Excel（v2 格式）",
            data=boh_bytes,
            file_name=boh_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.info("沒有產生任何建議名單。")

    if shortage_rows:
        st.markdown("#### ⚠️ 缺口提醒")
        short_df = pd.DataFrame(shortage_rows).groupby(["日期", "星期", "職務"], as_index=False)["缺口"].sum()
        st.dataframe(short_df, use_container_width=True, hide_index=True)
    else:
        st.success("✅ 本週模板在目前人池下可完成覆蓋，無缺口。")

st.markdown("---")
st.caption("此頁面目前為建議器（不直接寫入 shifts）。建議先檢視結果，再到 Shifts 頁面落地班表。")
